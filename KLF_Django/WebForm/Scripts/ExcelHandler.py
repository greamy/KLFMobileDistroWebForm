import openpyxl
import os #added by ben 11/10
from django.conf import settings

class ExcelFile:

	def __init__(self, fileName, headers, directory):
		self.name = fileName
		self.headers = headers
		self.directory = directory
		self.wb = self.generateFile()
	
	def generateFile(self):
		try:
			# Assuming 'ExcelDocs' is the subdirectory you want to save the Excel file in
			#directory = 'WebForm/ExcelDocs'
			if not os.path.exists(self.directory):
				os.makedirs(self.directory)

			file_path = os.path.join(self.directory, self.name)
			if os.path.exists(file_path):
				wb = openpyxl.load_workbook(file_path)
			else:
				wb = openpyxl.Workbook()
				ws = wb.active
				ws.append(self.headers)
				wb.save(file_path)
				wb = openpyxl.load_workbook(file_path)
		except FileNotFoundError:
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.append(self.headers)
			wb.save(file_path)
			wb = openpyxl.load_workbook(file_path)

		ws = wb.active
		ws.title = self.name.split(".")[0]
		return wb

	def addData(self, data):
		ws = self.wb.active
		# check if multiple rows in data input

		if isinstance(data[0], list):
			for row in data:
				ws.append(row)
		else:
			ws.append(data)


	def saveFile(self):
		file_path = os.path.join(settings.BASE_DIR, self.directory, self.name)
		print("Saving to: " + str(file_path))
		self.wb.save(file_path)





#Below is the original Generate Files definition
#Ben modified this definition on 11/17
